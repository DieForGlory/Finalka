# app/services/geocoding_service.py
import os
import csv
import time
import json  # <-- ДОБАВЛЕНО
from flask import current_app
from rapidfuzz import process, fuzz
from openpyxl.utils import column_index_from_string

# --- ИЗМЕНЕНИЕ: Импорт Redis ---
from app.extensions import redis_client

# Константа времени жизни ключа в Redis (24 часа)
TASK_EXPIRY_TIME_SECONDS = 86400
# --- КОНЕЦ ИЗМЕНЕНИЯ ---

_address_data = {}
_last_load_time = 0


# --- НОВАЯ ХЕЛПЕР-ФУНКЦИЯ ДЛЯ ОБНОВЛЕНИЯ СТАТУСА В REDIS ---
# (Скопирована из excel_processor.py для автономности)
def _update_task_status(task_id, status, progress=None, warnings_list=None, template_filename=None):
    """
    Безопасно обновляет статус задачи в Redis.
    """
    if not redis_client:
        print(f"[{task_id}] КРИТИКА: REDIS НЕ ДОСТУПЕН. Статус не обновлен.")
        return

    try:
        current_data_json = redis_client.get(task_id)
        if current_data_json:
            data = json.loads(current_data_json)
        else:
            data = {'owner_id': None}

        data['status'] = status
        if progress is not None:
            data['progress'] = progress
        if warnings_list is not None:
            data['warnings'] = warnings_list
        if template_filename is not None:
            data['template_filename'] = template_filename

        redis_client.setex(
            task_id,
            TASK_EXPIRY_TIME_SECONDS,
            json.dumps(data)
        )
    except Exception as e:
        print(f"[{task_id}] ОШИБКА: Не удалось обновить статус в Redis: {e}")


# --- КОНЕЦ ХЕЛПЕР-ФУНКЦИИ ---


def load_addresses(force=False):
    """
    Загружает адреса из CSV-файла в кэш.
    """
    global _address_data, _last_load_time
    file_path = current_app.config['ADDRESS_CSV_FILE']

    # Кэширование на 10 минут
    if not force and (time.time() - _last_load_time < 600):
        return

    if not os.path.exists(file_path):
        _address_data = {}
        print("[GeocodingService] Файл addresses.csv не найден.")
        return

    temp_data = {}
    try:
        with open(file_path, mode='r', encoding='utf-8') as f:
            reader = csv.reader(f)
            # Пропускаем заголовок
            try:
                next(reader)
            except StopIteration:
                pass  # Файл пуст

            for row in reader:
                if len(row) >= 3:
                    address = row[0].strip().lower()
                    try:
                        lat, lon = float(row[1]), float(row[2])
                        temp_data[address] = (lat, lon)
                    except (ValueError, TypeError):
                        pass  # Пропускаем строки с неверными координатами

        _address_data = temp_data
        _last_load_time = time.time()
        print(f"[GeocodingService] Загружено {_len_(_address_data)} адресов.")

    except Exception as e:
        print(f"[GeocodingService] Ошибка загрузки addresses.csv: {e}")


def _find_best_match(address):
    """
    Находит наилучшее совпадение адреса в загруженном кэше.
    """
    if not _address_data:
        load_addresses()
        if not _address_data:
            return None  # Кэш пуст или не загружен

    # Ищем совпадение
    # (limit=1 возвращает 1 самое похожее совпадение)
    best_match = process.extractOne(
        address.lower(),
        _address_data.keys(),
        scorer=fuzz.WRatio,
        score_cutoff=90  # Порог совпадения
    )

    if best_match:
        # best_match это кортеж (найденный_адрес, оценка, ключ)
        found_address_key = best_match[2]
        return _address_data[found_address_key]  # Возвращаем (lat, lon)

    return None


def apply_post_processing(task_id, template_wb, t_start_row, post_function):
    """
    Применяет функции пост-обработки (например, геокодинг) к файлу.
    """
    # --- ИЗМЕНЕНИЕ: 'task_statuses' удален из аргументов ---

    if post_function == 'none':
        return

    if post_function == 'geocode':
        print(f"[{task_id}] Запуск геокодинга...")

        # --- ИЗМЕНЕНИЕ: Обновляем статус через Redis ---
        _update_task_status(task_id, "Геокодирование...", 91)
        # --- КОНЕЦ ИЗМЕНЕНИЯ ---

        try:
            ws = template_wb.active
            max_row = ws.max_row
            if max_row <= t_start_row:
                return  # Нет данных

            # --- Находим колонки ---
            # (Предполагаем, что они называются "Адрес", "Широта", "Долгота"
            # и находятся в t_start_row)
            address_col, lat_col, lon_col = None, None, None
            for cell in ws[t_start_row]:
                val = str(cell.value).lower().strip()
                if val == 'адрес':
                    address_col = cell.column
                elif val == 'широта':
                    lat_col = cell.column
                elif val == 'долгота':
                    lon_col = cell.column

            if not all([address_col, lat_col, lon_col]):
                print(f"[{task_id}] Ошибка геокодинга: не найдены колонки 'Адрес', 'Широта', 'Долгота'.")
                # Обновляем статус, чтобы пользователь увидел ошибку
                _update_task_status(task_id, "Ошибка: не найдены колонки 'Адрес', 'Широта', 'Долгота'.", 92)
                return

            total_rows = max_row - t_start_row
            processed_rows = 0
            progress_step = 5  # (91% -> 96%)

            for row_idx in range(t_start_row + 1, max_row + 1):
                address_cell = ws.cell(row=row_idx, column=address_col)
                address = address_cell.value

                if address:
                    coords = _find_best_match(str(address))
                    if coords:
                        ws.cell(row=row_idx, column=lat_col).value = coords[0]
                        ws.cell(row=row_idx, column=lon_col).value = coords[1]

                processed_rows += 1

                if processed_rows % 50 == 0:  # Обновляем каждые 50 строк
                    # --- ИЗМЕНЕНИЕ: Обновляем статус через Redis ---
                    _update_task_status(
                        task_id,
                        f"Геокодирование... {processed_rows}/{total_rows}",
                        int(91 + (progress_step * (processed_rows / total_rows)))
                    )
                    # --- КОНЕЦ ИЗМЕНЕНИЯ ---

            print(f"[{task_id}] Геокодирование завершено.")
            _update_task_status(task_id, "Геокодирование завершено", 96)

        except Exception as e:
            print(f"[{task_id}] КРИТИЧЕСКАЯ ОШИБКА геокодинга: {e}")
            _update_task_status(task_id, f"Ошибка геокодинга: {e}", 95)

    else:
        print(f"[{task_id}] Неизвестная функция пост-обработки: {post_function}")